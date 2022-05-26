from selenium import webdriver
from selenium.webdriver.common.by import By

from tqdm import tqdm
from typing import Tuple

# pip install openpyxl
import openpyxl
import os

from utils import do_crawling

# 🌟 important parameter. 크롤링 할 페이지 지정
NUM_CRAWL_PAGE = 10

# 🚀 크롤링 완료된 정보를 저장할 excel sheet_codi 지정
wb_codi = openpyxl.Workbook()
sheet_codi = wb_codi.active
sheet_codi.append(["id",  "style", "img_url", "url"])

wb_codi_tag = openpyxl.Workbook()
sheet_codi_tag = wb_codi_tag.active
sheet_codi_tag.append(["id", "tag"])

wb_codi_item_id = openpyxl.Workbook()
sheet_codi_item_id = wb_codi_item_id.active
sheet_codi_item_id.append(["id", "item_id"])

workbooks = (wb_codi, wb_codi_tag, wb_codi_item_id)
sheets = (sheet_codi, sheet_codi_tag, sheet_codi_item_id)


# 🚀 지정한 페이지 수 만큼 크롤링 진행하기
do_crawling(workbooks, sheets, NUM_CRAWL_PAGE)