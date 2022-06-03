# pip install openpyxl
import openpyxl

from utils import do_crawling

# 🌟 important parameter. 크롤링 할 페이지 지정
NUM_CRAWL_PAGE = 1

# 🚀 크롤링 완료된 정보를 저장할 excel sheet_codi 지정
wb_codi = openpyxl.Workbook()
sheet_codi = wb_codi.active
sheet_codi.append(["id",  "style", "img_url", "url", "popularity"])

wb_codi_tag = openpyxl.Workbook()
sheet_codi_tag = wb_codi_tag.active
sheet_codi_tag.append(["id", "tag"])

wb_item_codi_id = openpyxl.Workbook()
sheet_item_codi_id = wb_item_codi_id.active
sheet_item_codi_id.append(["id", "codi_id"])

workbooks = (wb_codi, wb_codi_tag, wb_item_codi_id)
sheets = (sheet_codi, sheet_codi_tag, sheet_item_codi_id)


# 🚀 지정한 페이지 수 만큼 크롤링 진행하기
do_crawling(workbooks, sheets, NUM_CRAWL_PAGE)